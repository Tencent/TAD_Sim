#!/usr/bin/env python3
# -*- coding: utf-8 -*-
from __future__ import annotations

import argparse
import re
import shutil
import time
from dataclasses import dataclass, field
from pathlib import Path
from typing import Dict

import cv2
import easyocr
import numpy as np
from loguru import logger
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# import paddleocr


@dataclass(order=True)
class Base:
    def __post_init__(self) -> None:
        self.threshold = 0.7

    def set_log_file(self, pathdir: Path, filename_log: str = "file") -> None:
        """
        > Create log files

        Args:
          pathdir (Path): The path of the output directory
          filename_log (str): The name of the log file

        Returns:
          None
        """
        # 创建 log 文件
        logger.opt(lazy=True).add(pathdir / f"{filename_log}.log", rotation="100 MB", encoding="utf-8")

    def create_folder_structure(self, name: str) -> Path:
        """Create output folders"""
        # 创建 xosc xodr 子文件夹
        try:
            pathdir_output = Path(__file__).resolve().parent / f"{name}"

            # 创建输出文件夹(检查文件夹是否存在,存在则删除重新创建)
            if pathdir_output.is_dir():
                shutil.rmtree(pathdir_output)
            pathdir_output.mkdir(parents=True, exist_ok=True, mode=0o755)

            return pathdir_output

        except OSError as e:
            logger.opt(lazy=True).error(f"Error creating output directories: {e}")
            raise

    def event_occurred(self, gray_frame: np.ndarray, template: np.ndarray, name: str) -> bool:
        # 在视频帧中查找模板图片
        res = cv2.matchTemplate(gray_frame, template, cv2.TM_CCOEFF_NORMED)
        _, max_val, _, _ = cv2.minMaxLoc(res)

        # 比较当前帧和特定事件图片, 返回是否出现特定事件
        logger.opt(lazy=True).debug(f"查找模板图片{name}, max_val = {round(max_val, 3)}")

        return max_val >= self.threshold

    def get_event_templates(self, pathdir_template: Path) -> Dict[str, np.ndarray]:
        """
        > Get templates
        """
        out: Dict[str, np.ndarray] = {
            file.stem: cv2.imread(str(file), 0) for file in pathdir_template.glob("*") if file.is_file()
        }
        return out


@dataclass(order=True)
class VideoProcessor(Base):
    # 你的视频文件路径
    pathfile_video: Path
    # 场景结束文字模板图片路径
    pathdir_template: Path = field(default=Path("event_template_27"))

    def __post_init__(self) -> None:
        self.pathdir_output = self.create_folder_structure(self.pathfile_video.stem)
        self.set_log_file(pathdir=self.pathdir_output)

        # 提高处理速度
        self.skip_sec = 3  # 每隔多少秒处理一次

        self.threshold = 0.7

        # 场景名识别指定区域
        # ocr 不指定区域识别效果不好, 指定区域识别效果更好,
        # 但是对于录屏有要求,  27寸显示器, 分辨率 3840 * 2160
        # tadsim 全屏录制
        self.filename_roi = (353, 153, 2647, 77)  # (x, y, width, height)
        # 24 寸显示器
        # self.filename_roi = (233, 101, 800, 70)  # (x, y, width, height)
        # 联想 27 寸显示器
        # self.filename_roi = (240, 100, 1300, 170)  # (x, y, width, height)

        # # 场景名识别关键字
        # start_char = r"场景名: "
        # end_char = r"xosc|sim|\."
        # self.pat_find = re.compile(f"{start_char}(.*?)(?:{end_char})", re.IGNORECASE)

        # 场景名识别关键字
        start_char = r": "
        end_char = r"XOSC|xosc|sim|\.|tosc|O5C|OSC|xO5C|zosc|播放"
        # self.pat_find = re.compile(f"{start_char}(.*?)(?:{end_char})", re.IGNORECASE)
        self.pat_find = re.compile(f"{start_char}(.*){end_char}", re.IGNORECASE)

        #
        replacement_dict = {
            " ": "_",
            r"\.": "",
            "__": "_",
            "0g": "09",
            "ZO": "20",
            "Z1": "21",
            r"1Z5|1ZS|12S": "125",
            r"25O|Z5O|ZSO|Z50|ZS0": "250",
            r"SOO|5OO|S0O|SO0|5O0|50O": "500",
            "StopGO": "StopGo",
            "CCRCCRm": "CCR_CCRm",
            r"_1O|_1o": "_10",
            r"_1Z|_1z": "_12",
            r"_5S|_SS": "_55",
            r"_5Z|_SZ": "_52",
            r"_5O|_SO": "_50",
            r"_Z4": "_24",
            r"_Z5": "_25",
            r"_Z7": "_27",
            r"_ZG": "_29",
            r"_ZZ": "_22",
            r"LASO|LA5O": "LA50",
            r"FASO|FA5O": "FA50",
            r"G$|g$": "9",
            r"Z$|z$": "2",
            r"S$|s$": "5",
            r"O$|o$": "0",
            r"_#$|_$": "_1",
        }
        self.compiled_regex_dict = {re.compile(pattern): repl for pattern, repl in replacement_dict.items()}

    def _event_occurred_any(self, gray_frame: np.ndarray, templates: Dict[str, np.ndarray]) -> bool:
        return any(self.event_occurred(gray_frame, v, k) for k, v in templates.items())

    def _replace_with_dict(self, text: str) -> str:
        # print("input text: ", text)
        # 使用已编译的正则表达式对象对输入字符串进行多次替换
        for pattern, repl in self.compiled_regex_dict.items():
            # print(f"pattern: {pattern}, repl: {repl}")
            text = pattern.sub(repl, text)
            # print(f"output text: {text}")

        return text

    def _extract_between_chars(self, text: str) -> str:
        print(f"{text = }")

        outs = self.pat_find.findall(text)

        print(f"{outs = }")

        return self._replace_with_dict(outs[0]) if outs else ""

    def _ocr_find_filename(self, ocr, frame: np.ndarray):
        logger.opt(lazy=True).debug("使用 OCR 读取整个帧中的文字, 并找到需要的内容")

        # 从指定区域的文字中识别并提取特定字段
        x, y, w, h = self.filename_roi
        texts = ocr.readtext(frame[y : y + h, x : x + w], detail=0)
        # texts = ocr.ocr(frame[y : y + h, x : x + w], cls=True)  # paddleocr fail

        logger.opt(lazy=True).info(f"texts: {texts}")

        text_str = " ".join(texts).strip()

        return self._extract_between_chars(text_str)

    def _save_capture_frame(self, frame: np.ndarray, pathfile: Path) -> None:
        logger.opt(lazy=True).info(f"截取连续事件的第一帧并保存: {pathfile.name}")

        basename = pathfile.stem
        ext = pathfile.suffix
        counter = 1
        while pathfile.exists():
            pathfile = pathfile.parent / f"{basename}_{counter}{ext}"
            counter += 1
        cv2.imwrite(str(pathfile), frame)

    def process_video(self) -> None:
        # 初始化参数
        flag_event = False
        ocr = easyocr.Reader(lang_list=["en", "ch_sim"], gpu=True)
        # ocr = paddleocr.PaddleOCR(lang="ch")  # 初始化 PaddleOCR

        # 全部模板图片转换
        templates = self.get_event_templates(self.pathdir_template)

        # 加载视频
        cap = cv2.VideoCapture(str(self.pathfile_video))

        # 获取视频的帧率
        fps = cap.get(cv2.CAP_PROP_FPS)
        frame_count_skip = self.skip_sec * fps

        # while cap.isOpened():
        ret = cap.isOpened()
        while ret:
            # 获取当前视频帧位置, 基于以0开始的被捕获或解码的帧索引
            frame_count = cap.get(1)
            # 跳帧处理
            if frame_count % frame_count_skip != 0:
                ret = cap.grab()
                continue

            # 读取视频帧
            ret, frame = cap.read()

            # 确保图像被正确读取
            if frame is None:
                continue

            # 将视频帧转换为灰度图像
            gray_frame = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)

            # 当不是指定事件时, 重置 flag_event 标志, 并跳过当前帧
            if not self._event_occurred_any(gray_frame, templates):
                flag_event = False
                continue

            # 检测到是指定事件, 且 flag_event 为 False时
            # 从识别的文字中提取特定字段, 则将其用作截图文件名
            # 存储图片
            # 重置 flag_event 标志, 以便仅保存第一帧
            if not flag_event:
                filename = self._ocr_find_filename(ocr, frame)
                self._save_capture_frame(frame=frame, pathfile=self.pathdir_output / f"{filename}.png")
                flag_event = True

        # 释放视频资源
        cap.release()


@dataclass(order=True)
class ImageToExcel(Base):
    # 你的视频文件路径
    pathdir_image: Path
    suffix: str = field(default=".png")
    # pathdir_template: Path = field(default=Path("event_template"))

    def __post_init__(self) -> None:
        self.pathfile_xlsx = self.pathdir_image / "scene_check_report.xlsx"
        self.set_log_file(pathdir=self.pathdir_image)

        self.max_size = (300, 300)

        self.threshold = 0.7

        self.headers = ["场景名称", "期望场景表现", "实际场景表现", "验证图片"]

        self.expected_mapping = {
            "template_collision": "使用驾驶员模型匀速行驶, 与交通参与者发生碰撞",
            "template_endpoint": "正常行驶, 无碰撞发生",
            "template_maxtime": "正常行驶, 无碰撞发生",
            "none": "",
        }

        self.actual_mapping = {
            "template_collision": "发生碰撞",
            "template_endpoint": "正常结束, 无碰撞发生",
            "template_maxtime": "正常结束, 无碰撞发生",
            "none": "",
        }

    def _get_pathdir_template(self, name) -> Path:
        return Path(f"event_template_{name}")

    def get_event_name(self, gray_frame: np.ndarray, templates: Dict[str, np.ndarray]) -> str:
        return next((k for k, v in templates.items() if self.event_occurred(gray_frame, v, k)), "none")

    def set_cell_info(self, ws, row: int, col: int, value: str) -> None:
        cell = ws.cell(row=row, column=col)
        # 居中
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.value = value

        # 首行列填充背景色为 浅绿色
        if row == 1:
            cell.fill = PatternFill(start_color="C1FFC1", end_color="C1FFC1", fill_type="solid")
            cell.font = Font(bold=True)

    def save_one_sheet(self, subdir, templates, ws):
        # 获取图片列表
        png_files = list(subdir.glob(f"*{self.suffix}"))
        png_files.sort(key=lambda x: x.name)

        # 添加表头, 居中, 浅绿色背景
        for j, value in enumerate(self.headers, start=1):
            self.set_cell_info(ws, row=1, col=j, value=value)

        # 设置行高, 列宽
        ws.row_dimensions[1].height = 20
        ws.column_dimensions[get_column_letter(1)].width = 60
        ws.column_dimensions[get_column_letter(2)].width = 30
        ws.column_dimensions[get_column_letter(3)].width = 30
        ws.column_dimensions[get_column_letter(4)].width = 45

        for i, png_file in enumerate(png_files, start=2):
            # 设置单元格行高, 格式
            ws.row_dimensions[i].height = 150

            # 第1列逐行增加 场景名称
            self.set_cell_info(ws, row=i, col=1, value=f"{png_file.stem}.xosc")

            # 读取原始图片
            img = cv2.imread(str(png_file))

            # 第2列逐行
            gray_frame = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
            event_name = self.get_event_name(gray_frame, templates)

            self.set_cell_info(ws, row=i, col=2, value=self.expected_mapping[event_name])
            self.set_cell_info(ws, row=i, col=3, value=self.actual_mapping[event_name])

            # 第4列逐行增加图片, 将原始图片插入到 Excel 中, 调整图片在 Excel 中的显示大小
            xl_img = XLImage(str(png_file))
            h, w = img.shape[:2]
            scale = min(self.max_size[0] / w, self.max_size[1] / h)
            xl_img.width = int(w * scale)
            xl_img.height = int(h * scale)
            ws.add_image(xl_img, f"D{i}")

    def process(self):
        # 创建一个 Excel 工作簿
        wb = Workbook()
        # 遍历输入文件夹下的子文件夹
        for subdir in self.pathdir_image.iterdir():
            logger.opt(lazy=True).info(f"处理子文件夹: {subdir.name}")
            # print(subdir.name.split("_")[-1])
            if subdir.is_dir():
                # 使用子文件夹名称作为新工作表的名称
                ws = wb.create_sheet(subdir.name)

                pathdir_template = self._get_pathdir_template(subdir.name.split("_")[-1])

                templates = self.get_event_templates(pathdir_template)

                self.save_one_sheet(subdir, templates, ws)

        # 删除默认工作表
        default_sheet = wb["Sheet"]
        wb.remove(default_sheet)

        wb.save(str(self.pathfile_xlsx))


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--video", "-v", dest="pathfile_video", help="pathfile_video", metavar="FILE")
    parser.add_argument("--image", "-i", dest="pathdir_image", help="pathdir_image", metavar="FILE")
    args = parser.parse_args()

    t0 = time.time()
    if args.pathfile_video:
        vp = VideoProcessor(Path(args.pathfile_video))
        vp.process_video()
        t1 = time.time()
        logger.opt(lazy=True).info(f"处理视频耗时: {round(t1 - t0, 2)} s")

    elif args.pathdir_image:
        ie = ImageToExcel(Path(args.pathdir_image))
        ie.process()
        t1 = time.time()
        logger.opt(lazy=True).info(f"处理图片耗时: {round(t1 - t0, 2)} s")

    else:
        parser.print_help()


if __name__ == "__main__":
    main()
